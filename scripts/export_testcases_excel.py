from __future__ import annotations

import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".tools" / "openpyxl"))

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "docs/testcases/TC_JackFruit_PoolJackpot_UC01.md"
OUTPUT_DIR = ROOT / "outputs/jackfruit_pool_jackpot"
OUTPUT = OUTPUT_DIR / "TC_JackFruit_PoolJackpot_UC01-UC04.xlsx"


def clean_cell(value: str) -> str:
    value = value.strip().replace("<br>", "\n")
    value = re.sub(r"`([^`]*)`", r"\1", value)
    return value.replace("\\_", "_")


def markdown_table(text: str, heading: str) -> list[list[str]]:
    start = text.index(heading)
    lines = text[start:].splitlines()[1:]
    table_lines: list[str] = []
    in_table = False
    for line in lines:
        if line.startswith("|"):
            in_table = True
            table_lines.append(line)
        elif in_table:
            break
    rows = [[clean_cell(c) for c in line.strip().strip("|").split("|")] for line in table_lines]
    if len(rows) < 2:
        raise ValueError(f"Không tìm thấy bảng cho heading: {heading}")
    return [rows[0], *rows[2:]]


def open_questions(text: str) -> list[str]:
    start = text.index("# Open Questions")
    end = text.index("# Statistics", start)
    return [line[2:].strip() for line in text[start:end].splitlines() if line.startswith("- ")]


def set_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(vertical="center", shrink_to_fit=True)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=9, italic=True, color="44546A")
    ws["A2"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A2"].alignment = Alignment(vertical="center", shrink_to_fit=True)
    ws.row_dimensions[2].height = 22


def write_table(ws, title: str, subtitle: str, rows: list[list[str]], table_name: str, widths: list[int]) -> None:
    set_title(ws, title, subtitle, len(rows[0]))
    for col_idx, value in enumerate(rows[0], 1):
        cell = ws.cell(4, col_idx, value)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 32
    for row_idx, row in enumerate(rows[1:], 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = Font(name="Aptos", size=9, color="1F1F1F")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 60 if len(rows[0]) > 8 else 32
    last_row = 4 + len(rows) - 1
    last_col_letter = ws.cell(4, len(rows[0])).column_letter
    tab = Table(displayName=table_name, ref=f"A4:{last_col_letter}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A4:{last_col_letter}{last_row}"
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(4, idx).column_letter].width = width
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build() -> Path:
    text = SOURCE.read_text(encoding="utf-8")
    requirements = markdown_table(text, "# Tổng quan Requirement")
    risks = markdown_table(text, "# Đánh giá Risk")
    traceability = markdown_table(text, "# Traceability Matrix")
    test_cases = markdown_table(text, "# Test Case")
    questions = open_questions(text)

    if len(test_cases) - 1 != 85 or len(requirements) - 1 != 25:
        raise ValueError(f"Sai số lượng nguồn: {len(test_cases)-1} TC, {len(requirements)-1} requirements")

    wb = Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("Summary")
    tc_ws = wb.create_sheet("Test Cases")
    req_ws = wb.create_sheet("Requirements")
    risk_ws = wb.create_sheet("Risk Assessment")
    trace_ws = wb.create_sheet("Traceability")
    oq_ws = wb.create_sheet("Open Questions")

    set_title(summary, "POOL JACKPOT — TEST CASES", "UC-01 đến UC-04 • 85 Test Cases • Version 1.3", 6)
    summary["A4"] = "Chỉ số"
    summary["B4"] = "Giá trị"
    summary["D4"] = "Trạng thái thực thi"
    summary["E4"] = "Số lượng"
    for cell in (summary["A4"], summary["B4"], summary["D4"], summary["E4"]):
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center")
    high_risk_count = sum(1 for row in test_cases[1:] if row[4] == "High")
    critical_count = sum(1 for row in test_cases[1:] if row[3] == "Critical")
    metrics = [
        ("Tổng Test Case", len(test_cases) - 1),
        ("Tổng Requirement", len(requirements) - 1),
        ("High-Risk Test Cases", high_risk_count),
        ("Critical-Priority Test Cases", critical_count),
        ("Coverage", "100%"),
        ("Quality Score", "97/100"),
        ("Review Decision", "Approved"),
    ]
    for row, (label, value) in enumerate(metrics, 5):
        summary.cell(row, 1, label)
        if isinstance(value, str) and value.startswith("="):
            summary.cell(row, 2, value)
        else:
            summary.cell(row, 2, value)
    statuses = ["Not Run", "Pass", "Fail", "Blocked", "N/A"]
    for row, status in enumerate(statuses, 5):
        summary.cell(row, 4, status)
        summary.cell(row, 5, f'=COUNTIF(\'Test Cases\'!L5:L89,D{row})')
    summary["D11"] = "Pass Rate"
    summary["E11"] = '=IFERROR(E6/SUM(E6:E8),0)'
    summary["E11"].number_format = "0.0%"
    for row in range(5, 12):
        for col in (1, 2, 4, 5):
            summary.cell(row, col).fill = PatternFill("solid", fgColor="F3F6FA" if row % 2 else "E8EFF7")
            summary.cell(row, col).font = Font(name="Aptos", size=10)
            summary.cell(row, col).alignment = Alignment(vertical="center")
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 4
    summary.column_dimensions["D"].width = 22
    summary.column_dimensions["E"].width = 16
    summary.column_dimensions["F"].width = 4
    summary.freeze_panes = "A4"
    summary.sheet_view.showGridLines = False

    write_table(
        tc_ws,
        "TEST CASES — POOL JACKPOT",
        "85 test cases • UC-01 đến UC-04 • Cập nhật Status trực tiếp tại cột L",
        test_cases,
        "TestCasesTable",
        [12, 22, 24, 12, 10, 38, 45, 38, 55, 55, 18, 14, 30],
    )
    tc_last = 4 + len(test_cases) - 1
    status_validation = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked,N/A"', allow_blank=True)
    status_validation.error = "Chọn một trạng thái trong danh sách."
    status_validation.errorTitle = "Trạng thái không hợp lệ"
    tc_ws.add_data_validation(status_validation)
    status_validation.add(f"L5:L{tc_last}")
    status_colors = {"Pass": "C6EFCE", "Fail": "FFC7CE", "Blocked": "FFE699", "Not Run": "D9EAF7", "N/A": "E7E6E6"}
    for status, color in status_colors.items():
        tc_ws.conditional_formatting.add(
            f"L5:L{tc_last}",
            FormulaRule(formula=[f'$L5="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )
    tc_ws.auto_filter.ref = f"A4:M{tc_last}"

    write_table(req_ws, "REQUIREMENTS", "25 requirements trích xuất từ UC-01 đến UC-04", requirements, "RequirementsTable", [16, 110])
    write_table(risk_ws, "RISK ASSESSMENT", "Đánh giá theo Business Impact, Usage Frequency, Technical Complexity, Change Frequency và Historical Defect Rate", risks, "RiskTable", [16, 14, 14, 14, 105])
    write_table(trace_ws, "TRACEABILITY MATRIX", "Liên kết Requirement ↔ Test Case", traceability, "TraceabilityTable", [18, 100])
    oq_rows = [["STT", "Open Question"], *[[idx, question] for idx, question in enumerate(questions, 1)]]
    write_table(oq_ws, "OPEN QUESTIONS", "Các nội dung cần BA/PO/Dev xác nhận; không tự suy diễn", oq_rows, "OpenQuestionsTable", [10, 120])

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    assert check.sheetnames == ["Summary", "Test Cases", "Requirements", "Risk Assessment", "Traceability", "Open Questions"]
    assert check["Test Cases"].max_row == 89
    assert check["Requirements"].max_row == 29
    assert check["Test Cases"]["A5"].value == "TC-001"
    assert check["Test Cases"]["A89"].value == "TC-085"
    assert check["Summary"]["B5"].value == 85
    check.close()
    return OUTPUT


if __name__ == "__main__":
    print(build())
