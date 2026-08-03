from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

source = Path("docs/testcases/TC_Danh_sach_Tenant.md")
target = Path("docs/testcases/TC_Danh_sach_Tenant.xlsx")
lines = source.read_text(encoding="utf-8").splitlines()
start = lines.index("## Test Case") + 1
table_lines = []
for line in lines[start:]:
    if line.startswith("## "):
        break
    if line.startswith("|"):
        table_lines.append(line)

def cells(line):
    return [cell.strip().replace("<br>", "\n") for cell in line.strip().strip("|").split("|")]

headers = cells(table_lines[0])
rows = [cells(line) for line in table_lines[2:]]
book = Workbook()
sheet = book.active
sheet.title = "Danh sách Tenant"
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = f"A1:M{len(rows) + 1}"
header_fill = PatternFill("solid", fgColor="E8EEF5")
for col, value in enumerate(headers, 1):
    cell = sheet.cell(1, col, value)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row_index, row in enumerate(rows, 2):
    for col, value in enumerate(row, 1):
        cell = sheet.cell(row_index, col, value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
sheet.row_dimensions[1].height = 30
for row_index in range(2, len(rows) + 2):
    sheet.row_dimensions[row_index].height = 100
widths = [12, 15, 18, 10, 12, 40, 35, 30, 58, 68, 20, 12, 35]
for index, width in enumerate(widths, 1):
    sheet.column_dimensions[chr(64 + index)].width = width
book.save(target)
print(target.resolve())
